"""
预处理：从 WoS Excel 提取原始字段，输出纯文本 TSV。
不做任何清洗、分词或映射——只提取原始文本。
"""

import os
import openpyxl

INPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "0325")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "chunks")
YEARS = list(range(2012, 2019))  # 2012-2018
SHEET_NAMES = ["ontology", "KG", "LinkedData", "Thesaurus"]

# 从 country_code 文件中提取已知国家名，用于内容匹配兜底
KNOWN_COUNTRIES = {
    "Peoples R China", "United States", "Germany", "France", "England",
    "Spain", "Italy", "Canada", "Australia", "Japan", "South Korea",
    "Brazil", "India", "Netherlands", "Switzerland", "Sweden", "Belgium",
    "Taiwan", "Portugal", "Poland", "Austria", "Greece", "Czech Republic",
    "Turkey", "Iran", "Russia", "Norway", "Finland", "Denmark", "Ireland",
    "Scotland", "Wales", "North Ireland", "Singapore", "South Africa",
    "New Zealand", "Mexico", "Argentina", "Chile", "Colombia",
    "Malaysia", "Thailand", "Indonesia", "Vietnam", "Pakistan",
    "Saudi Arabia", "Egypt", "Nigeria", "Kenya", "Morocco", "Tunisia",
    "Romania", "Hungary", "Bulgaria", "Croatia", "Slovenia", "Slovakia",
    "Serbia", "Ukraine", "Israel", "United Arab Emirates", "Qatar",
    "Estonia", "Latvia", "Lithuania", "Cyprus", "Luxembourg",
    "Algeria", "Bangladesh", "Sri Lanka", "Philippines", "Peru",
    "Jordan", "Oman", "Kazakhstan", "Venezuela", "Ecuador", "Jamaica",
}

def guess_column_index(header_row):
    """根据列名关键词定位列索引。每个字段只匹配第一个命中，防止被后续相似列覆盖。"""
    mapping = {}
    for col_idx, cell in enumerate(header_row):
        val = str(cell.value).strip() if cell.value else ""
        if not val:
            continue
        vl = val.lower()
        # Year: 精确匹配
        if vl == "year" and "Year" not in mapping:
            mapping["Year"] = col_idx
        # Keywords: 精确匹配 "Keywords"，不匹配 "Keywords Plus"
        elif vl == "keywords" and "Keywords" not in mapping:
            mapping["Keywords"] = col_idx
        # Categories: 同时包含 categories 和 classification，排除 "Web of Science Categories" 等
        elif ("categories" in vl and "classification" in vl) and "Categories" not in mapping:
            mapping["Categories"] = col_idx
        elif vl == "country" and "country" not in mapping:
            mapping["country"] = col_idx
        elif vl == "country_code" and "country_code" not in mapping:
            mapping["country_code"] = col_idx
        elif vl == "research areas" and "Research Areas" not in mapping:
            mapping["Research Areas"] = col_idx
    return mapping


def detect_by_content(ws, data_start):
    """扫描前 5 行数据，按内容模式定位列。表头损坏时的兜底方案。"""
    sample = []
    for row in ws.iter_rows(
        min_row=data_start,
        max_row=min(data_start + 4, ws.max_row),
        values_only=True,
    ):
        sample.append(row)
    if not sample:
        return {}

    ncols = max(len(r) for r in sample)
    mapping = {}

    for ci in range(ncols):
        vals = []
        for r in sample:
            v = str(r[ci]) if ci < len(r) and r[ci] else ""
            vals.append(v)

        # Year 列：全为数字年份
        if all(v.isdigit() and 2000 <= int(v) <= 2030 for v in vals if v):
            mapping["Year"] = ci

        # Keywords 列：以 "Author Keywords" 开头
        if sum(1 for v in vals if v.startswith("Author Keywords")) >= len(vals) * 0.5:
            mapping["Keywords"] = ci

        # Categories 列：包含 "Research Areas"
        if sum(1 for v in vals if "Research Areas" in v) >= len(vals) * 0.5:
            mapping["Categories"] = ci

        # country 列: 值匹配已知国家名
        country_matches = sum(1 for v in vals if v in KNOWN_COUNTRIES)
        if country_matches >= 2:
            mapping["country"] = ci

        # country_code 列：2 字母大写，在靠后 1/3 位置
        if ci > ncols * 0.6:
            codes = [v for v in vals if len(v) == 2 and v.isupper() and v.isalpha()]
            if len(codes) >= 3:
                mapping["country_code"] = ci

    return mapping


def extract_sheet(ws, year, sheet_name):
    """从单个 sheet 提取所有行，返回 list of dict。"""
    rows = []
    header = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))[0]
    col_map = guess_column_index(header)
    data_start = 2

    # 检查关键列是否缺失
    essential = ["Year", "Keywords", "country"]
    missing = [f for f in essential if f not in col_map]
    if missing:
        content_map = detect_by_content(ws, data_start)
        for f in missing:
            if f in content_map:
                col_map[f] = content_map[f]
                print(f"  [INFO] {year} {sheet_name}: col '{f}' detected by content at index {content_map[f]}")

    # 如果 country 依然缺失，用 country_code 位置反推
    if "country" not in col_map and "country_code" in col_map:
        cc_idx = col_map["country_code"]
        if cc_idx > 0:
            col_map["country"] = cc_idx - 1
            print(f"  [INFO] {year} {sheet_name}: 'country' inferred at col {cc_idx - 1} (before country_code)")

    # 最后兜底：已知的硬编码位置（仅 2018 KG）
    if "country" not in col_map and year == 2018 and sheet_name == "KG":
        col_map["country"] = 21
        print(f"  [INFO] {year} {sheet_name}: 'country' using hardcoded fallback col 21")

    # 仍然缺失则报错
    still_missing = [f for f in essential if f not in col_map]
    if still_missing:
        print(f"  [ERROR] {year} {sheet_name}: cannot locate columns: {still_missing}")
        return rows

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if all(v is None for v in row):
            continue

        def get_val(field):
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx]
            return str(v).strip() if v is not None else ""

        year_val = get_val("Year")
        if not year_val:
            continue

        keywords_raw = get_val("Keywords")
        categories_raw = get_val("Categories")
        research_areas_raw = get_val("Research Areas")
        country = get_val("country")
        country_code = get_val("country_code")

        # 研究领域：优先 Research Areas 列，否则 Categories 列
        raw_research = research_areas_raw if research_areas_raw else categories_raw

        rows.append({
            "year": year_val,
            "keywords_raw": keywords_raw,
            "research_raw": raw_research,
            "country": country,
            "country_code": country_code,
            "sheet": sheet_name,
        })

    return rows


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    all_rows = []
    stats = {}

    for year in YEARS:
        fname = f"四张sheet_国家_颜色{year}.xlsx"
        fpath = os.path.join(INPUT_DIR, fname)
        if not os.path.exists(fpath):
            print(f"[SKIP] {fpath} not found")
            continue

        print(f"Processing: {fname}")
        wb = openpyxl.load_workbook(fpath, data_only=True)
        year_total = 0

        for sname in SHEET_NAMES:
            if sname not in wb.sheetnames:
                print(f"  [SKIP] Sheet {sname} not found")
                continue
            ws = wb[sname]
            rows = extract_sheet(ws, year, sname)
            count = len(rows)
            year_total += count
            stats[f"{year}_{sname}"] = count
            print(f"  {sname}: {count} rows")
            all_rows.extend(rows)

        wb.close()
        print(f"  -> {year} total: {year_total} rows\n")

    # 写出 all_rows.tsv（转义换行符和制表符以保证每行一条记录）
    tsv_path = os.path.join(OUTPUT_DIR, "all_rows.tsv")
    with open(tsv_path, "w", encoding="utf-8") as f:
        f.write("year\tkeywords_raw\tresearch_raw\tcountry\tcountry_code\tsheet\n")
        for r in all_rows:
            vals = [
                str(r["year"]),
                r["keywords_raw"].replace("\n", " ").replace("\r", " ").replace("\t", " "),
                r["research_raw"].replace("\n", " ").replace("\r", " ").replace("\t", " "),
                r["country"].replace("\n", " ").replace("\r", " ").replace("\t", " "),
                r["country_code"].replace("\n", " ").replace("\r", " ").replace("\t", " "),
                r["sheet"],
            ]
            f.write("\t".join(vals) + "\n")

    print(f"TSV written: {tsv_path}")
    print(f"Total rows: {len(all_rows)}")

    # 统计
    print("\n=== 按年份/Sheet 统计 ===")
    for year in YEARS:
        yt = sum(stats.get(f"{year}_{s}", 0) for s in SHEET_NAMES)
        if yt == 0:
            continue
        parts = " | ".join(f"{s}={stats.get(f'{year}_{s}', 0)}" for s in SHEET_NAMES)
        print(f"  {year}: {yt} ({parts})")


if __name__ == "__main__":
    main()
