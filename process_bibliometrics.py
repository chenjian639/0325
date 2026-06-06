"""
Bibliometric Data Processing Pipeline
Processes 2012-2018 Excel files to extract keyword frequencies
grouped by research field, country, and year.
"""

import argparse
import os
import re
import sys
from collections import defaultdict, Counter
from dataclasses import dataclass, field

import openpyxl
import pandas as pd

# ── Configuration ──────────────────────────────────────────────────────────

INPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "0325")
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
KEYWORD_DICT_PATH = os.path.join(OUTPUT_DIR, "keyword_dictionary.txt")
YEARS = list(range(2012, 2019))  # 2012-2018
SHEET_NAMES = ["ontology", "KG", "LinkedData", "Thesaurus"]

# ── WoS Research Areas → Broad Categories (Zendesk official mapping) ──

RESEARCH_AREA_TO_BROAD = {
    # Arts & Humanities (15)
    "Architecture": "Arts & Humanities",
    "Art": "Arts & Humanities",
    "Arts & Humanities Other Topics": "Arts & Humanities",
    "Asian Studies": "Arts & Humanities",
    "Classics": "Arts & Humanities",
    "Dance": "Arts & Humanities",
    "Film, Radio & Television": "Arts & Humanities",
    "History": "Arts & Humanities",
    "History & Philosophy of Science": "Arts & Humanities",
    "Literature": "Arts & Humanities",
    "Music": "Arts & Humanities",
    "Philosophy": "Arts & Humanities",
    "Religion": "Arts & Humanities",
    "Theater": "Arts & Humanities",

    # Life Sciences & Biomedicine (76)
    "Agriculture": "Life Sciences & Biomedicine",
    "Allergy": "Life Sciences & Biomedicine",
    "Anatomy & Morphology": "Life Sciences & Biomedicine",
    "Anesthesiology": "Life Sciences & Biomedicine",
    "Anthropology": "Life Sciences & Biomedicine",
    "Audiology & Speech-Language Pathology": "Life Sciences & Biomedicine",
    "Behavioral Sciences": "Life Sciences & Biomedicine",
    "Biochemistry & Molecular Biology": "Life Sciences & Biomedicine",
    "Biodiversity & Conservation": "Life Sciences & Biomedicine",
    "Biophysics": "Life Sciences & Biomedicine",
    "Biotechnology & Applied Microbiology": "Life Sciences & Biomedicine",
    "Cardiovascular System & Cardiology": "Life Sciences & Biomedicine",
    "Cell Biology": "Life Sciences & Biomedicine",
    "Critical Care Medicine": "Life Sciences & Biomedicine",
    "Dentistry, Oral Surgery & Medicine": "Life Sciences & Biomedicine",
    "Dermatology": "Life Sciences & Biomedicine",
    "Developmental Biology": "Life Sciences & Biomedicine",
    "Emergency Medicine": "Life Sciences & Biomedicine",
    "Endocrinology & Metabolism": "Life Sciences & Biomedicine",
    "Entomology": "Life Sciences & Biomedicine",
    "Environmental Sciences & Ecology": "Life Sciences & Biomedicine",
    "Evolutionary Biology": "Life Sciences & Biomedicine",
    "Fisheries": "Life Sciences & Biomedicine",
    "Food Science & Technology": "Life Sciences & Biomedicine",
    "Forestry": "Life Sciences & Biomedicine",
    "Gastroenterology & Hepatology": "Life Sciences & Biomedicine",
    "General & Internal Medicine": "Life Sciences & Biomedicine",
    "Genetics & Heredity": "Life Sciences & Biomedicine",
    "Geriatrics & Gerontology": "Life Sciences & Biomedicine",
    "Health Care Sciences & Services": "Life Sciences & Biomedicine",
    "Hematology": "Life Sciences & Biomedicine",
    "Immunology": "Life Sciences & Biomedicine",
    "Infectious Diseases": "Life Sciences & Biomedicine",
    "Integrative & Complementary Medicine": "Life Sciences & Biomedicine",
    "Legal Medicine": "Life Sciences & Biomedicine",
    "Life Sciences Biomedicine Other Topics": "Life Sciences & Biomedicine",
    "Marine & Freshwater Biology": "Life Sciences & Biomedicine",
    "Mathematical & Computational Biology": "Life Sciences & Biomedicine",
    "Medical Ethics": "Life Sciences & Biomedicine",
    "Medical Informatics": "Life Sciences & Biomedicine",
    "Medical Laboratory Technology": "Life Sciences & Biomedicine",
    "Microbiology": "Life Sciences & Biomedicine",
    "Mycology": "Life Sciences & Biomedicine",
    "Neurosciences & Neurology": "Life Sciences & Biomedicine",
    "Nursing": "Life Sciences & Biomedicine",
    "Nutrition & Dietetics": "Life Sciences & Biomedicine",
    "Obstetrics & Gynecology": "Life Sciences & Biomedicine",
    "Oncology": "Life Sciences & Biomedicine",
    "Ophthalmology": "Life Sciences & Biomedicine",
    "Orthopedics": "Life Sciences & Biomedicine",
    "Otorhinolaryngology": "Life Sciences & Biomedicine",
    "Paleontology": "Life Sciences & Biomedicine",
    "Parasitology": "Life Sciences & Biomedicine",
    "Pathology": "Life Sciences & Biomedicine",
    "Pediatrics": "Life Sciences & Biomedicine",
    "Pharmacology & Pharmacy": "Life Sciences & Biomedicine",
    "Physiology": "Life Sciences & Biomedicine",
    "Plant Sciences": "Life Sciences & Biomedicine",
    "Psychiatry": "Life Sciences & Biomedicine",
    "Public, Environmental & Occupational Health": "Life Sciences & Biomedicine",
    "Radiology, Nuclear Medicine & Medical Imaging": "Life Sciences & Biomedicine",
    "Rehabilitation": "Life Sciences & Biomedicine",
    "Reproductive Biology": "Life Sciences & Biomedicine",
    "Research & Experimental Medicine": "Life Sciences & Biomedicine",
    "Respiratory System": "Life Sciences & Biomedicine",
    "Rheumatology": "Life Sciences & Biomedicine",
    "Sport Sciences": "Life Sciences & Biomedicine",
    "Substance Abuse": "Life Sciences & Biomedicine",
    "Surgery": "Life Sciences & Biomedicine",
    "Toxicology": "Life Sciences & Biomedicine",
    "Transplantation": "Life Sciences & Biomedicine",
    "Tropical Medicine": "Life Sciences & Biomedicine",
    "Urology & Nephrology": "Life Sciences & Biomedicine",
    "Veterinary Sciences": "Life Sciences & Biomedicine",
    "Virology": "Life Sciences & Biomedicine",
    "Zoology": "Life Sciences & Biomedicine",

    # Physical Sciences (17)
    "Astronomy & Astrophysics": "Physical Sciences",
    "Chemistry": "Physical Sciences",
    "Crystallography": "Physical Sciences",
    "Electrochemistry": "Physical Sciences",
    "Geochemistry & Geophysics": "Physical Sciences",
    "Geology": "Physical Sciences",
    "Mathematics": "Physical Sciences",
    "Meteorology & Atmospheric Sciences": "Physical Sciences",
    "Mineralogy": "Physical Sciences",
    "Mining & Mineral Processing": "Physical Sciences",
    "Oceanography": "Physical Sciences",
    "Optics": "Physical Sciences",
    "Physical Geography": "Physical Sciences",
    "Physics": "Physical Sciences",
    "Polymer Science": "Physical Sciences",
    "Thermodynamics": "Physical Sciences",
    "Water Resources": "Physical Sciences",

    # Social Sciences (25)
    "Archaeology": "Social Sciences",
    "Area Studies": "Social Sciences",
    "Biomedical Social Sciences": "Social Sciences",
    "Business & Economics": "Social Sciences",
    "Communication": "Social Sciences",
    "Criminology & Penology": "Social Sciences",
    "Cultural Studies": "Social Sciences",
    "Demography": "Social Sciences",
    "Development Studies": "Social Sciences",
    "Education & Educational Research": "Social Sciences",
    "Ethnic Studies": "Social Sciences",
    "Family Studies": "Social Sciences",
    "Geography": "Social Sciences",
    "Government & Law": "Social Sciences",
    "International Relations": "Social Sciences",
    "Linguistics": "Social Sciences",
    "Mathematical Methods In Social Sciences": "Social Sciences",
    "Psychology": "Social Sciences",
    "Public Administration": "Social Sciences",
    "Social Issues": "Social Sciences",
    "Social Sciences Other Topics": "Social Sciences",
    "Social Work": "Social Sciences",
    "Sociology": "Social Sciences",
    "Urban Studies": "Social Sciences",
    "Women's Studies": "Social Sciences",

    # Technology (21)
    "Acoustics": "Technology",
    "Automation & Control Systems": "Technology",
    "Computer Science": "Technology",
    "Construction & Building Technology": "Technology",
    "Energy & Fuels": "Technology",
    "Engineering": "Technology",
    "Imaging Science & Photographic Technology": "Technology",
    "Information Science & Library Science": "Technology",
    "Instruments & Instrumentation": "Technology",
    "Materials Science": "Technology",
    "Mechanics": "Technology",
    "Metallurgy & Metallurgical Engineering": "Technology",
    "Microscopy": "Technology",
    "Nuclear Science & Technology": "Technology",
    "Operations Research & Management Science": "Technology",
    "Remote Sensing": "Technology",
    "Robotics": "Technology",
    "Science & Technology Other Topics": "Technology",
    "Spectroscopy": "Technology",
    "Telecommunications": "Technology",
    "Transportation": "Technology",
}

# Flat list used by extract_research_areas (greedy match)
KNOWN_RESEARCH_AREAS = list(RESEARCH_AREA_TO_BROAD.keys())

# Positional fallback column indices per sheet type (0-indexed)
FALLBACK_COLUMNS = {
    "ontology": {"Keywords": 10, "Categories": 12, "country": 21, "country_code": 42},
    "KG": {"Keywords": 10, "Categories": 12, "country": 21, "country_code": 34},
    "LinkedData": {"Keywords": 10, "Categories": 12, "country": 21, "country_code": 42},
    "Thesaurus": {"Keywords": 10, "Categories": 12, "country": 23, "country_code": 33},
}

# ── Data Structures ─────────────────────────────────────────────────────────

@dataclass
class ErrorRecord:
    source_year: int
    source_sheet: str
    source_row: int
    reason: str
    raw_keywords: str = ""
    raw_categories: str = ""
    raw_country: str = ""


# ── Column Mapping ──────────────────────────────────────────────────────────

def build_column_index(header_row, sheet_name, year):
    """Build {key: col_index} from header row, with positional fallback."""
    mapping = {}
    for col_idx, cell_value in enumerate(header_row):
        if cell_value is None:
            continue
        cv = str(cell_value).strip()
        if cv == "Year":
            mapping["Year"] = col_idx
        elif cv == "Keywords":
            mapping["Keywords"] = col_idx
        elif cv == "country":
            mapping["country"] = col_idx
        elif cv == "country_code":
            mapping["country_code"] = col_idx
        elif "Categories" in cv and "Classification" in cv:
            mapping["Categories"] = col_idx

    # Apply positional fallback for any missing keys
    defaults = FALLBACK_COLUMNS.get(sheet_name, {})
    for key, idx in defaults.items():
        if key not in mapping and key in defaults:
            mapping[key] = idx
            print(f"  [WARN] {year} {sheet_name}: '{key}' not found in header, using positional fallback col {idx}")

    return mapping


# ── Keyword Parsing ─────────────────────────────────────────────────────────

def split_author_vs_kwplus(raw_str):
    """Split raw Keywords string into (author_kw_str_or_None, kwplus_str_or_None)."""
    if not isinstance(raw_str, str) or not raw_str.strip():
        return None, None

    s = raw_str.strip()

    has_author = "Author Keywords" in s
    has_kwplus = "Keywords Plus" in s

    if not has_author and not has_kwplus:
        return s, None

    author_part = None
    kwplus_part = None

    if has_kwplus:
        idx = s.index("Keywords Plus")
        kwplus_part = s[idx + len("Keywords Plus"):]
        before = s[:idx]
        if "Author Keywords" in before:
            aidx = before.index("Author Keywords") + len("Author Keywords")
            author_part = before[aidx:]
        else:
            author_part = before.strip() or None
    elif has_author:
        idx = s.index("Author Keywords") + len("Author Keywords")
        author_part = s[idx:]

    return author_part, kwplus_part


# Patterns for annotation labels (order matters: longer patterns first)
_LABEL_PATTERNS = [
    re.compile(r'From\s+[A-Za-z\s&]+?\s*Thesaurus\s*:?\s*'),
    re.compile(r'\bThesaurus\s*:?\s*', re.IGNORECASE),
    re.compile(r"\bAuthor'?s?\s*keywords?\s*:?\s*", re.IGNORECASE),
    re.compile(r'\bAuthors?\s*:?\s*', re.IGNORECASE),
    re.compile(r'\bOther\s*:?\s*', re.IGNORECASE),
    re.compile(r'\bRegional\s+terms?\s*:?\s*', re.IGNORECASE),
]


def split_author_keywords(text):
    """Split author keywords from concatenated WoS format.

    Handles:
    - CamelCase concatenation (lowercase->uppercase boundary)
    - Semicolons as delimiters
    - Thesaurus:/Author:/Authors: annotation labels
    - Period as section separator between labeled groups
    """
    if not text or not text.strip():
        return []

    text = text.strip()

    # Step 0: Strip annotation labels
    for pat in _LABEL_PATTERNS:
        text = pat.sub(' ', text)

    # Clean up: collapse whitespace, remove stray periods
    text = re.sub(r'\s+', ' ', text).strip()
    text = re.sub(r'\s*\.\s*$', '', text)  # trailing period

    # Step 1: Split on semicolons
    parts = text.split(';')
    all_tokens = []

    for part in parts:
        part = part.strip()
        if not part:
            continue

        # Step 2: Split on period+space (section boundaries)
        subs = re.split(r'\.\s+', part)
        for sub in subs:
            sub = sub.strip(' .')
            if not sub:
                continue

            # Step 3: CamelCase boundary detection
            chars = []
            for i, ch in enumerate(sub):
                chars.append(ch)
                if (i < len(sub) - 1
                        and ch.islower()
                        and sub[i + 1].isupper()
                        and not ch.isspace()
                        and not sub[i + 1].isspace()):
                    chars.append('\x1F')

            tokens = ''.join(chars).split('\x1F')
            tokens = [t.strip().rstrip('.,;:!?') for t in tokens if t.strip()]
            all_tokens.extend(tokens)

    # Step 4: For tokens containing commas, try splitting if it looks like a list
    final_tokens = []
    for t in all_tokens:
        if ', ' in t and len(t) > 30:
            comma_parts = [p.strip() for p in t.split(', ') if p.strip()]
            final_tokens.extend(comma_parts)
        else:
            final_tokens.append(t)

    return final_tokens


def split_keywords_plus(text):
    """Split Keywords Plus text into individual terms.

    KeyWords Plus entries are concatenated without delimiters.
    Multi-word entries (e.g., 'RELATION EXTRACTION') have internal spaces.
    Different entries concatenated without spaces create long runs of alpha chars.

    Strategy: split on spaces first, producing alpha runs that are one or more
    concatenated entries. Long runs are re-split in Pass 2 using the dictionary.
    """
    if not text or not text.strip():
        return []

    # First, strip non-alpha, non-space junk
    cleaned = ''.join(c for c in text if c.isalpha() or c.isspace())
    # Split on whitespace to get individual alpha runs
    tokens = cleaned.split()
    return [t.strip().rstrip('.,;:!?') for t in tokens if t.strip()]


def normalize_token(token):
    """Clean up a parsed keyword token."""
    t = token.strip()
    # Collapse multiple spaces
    t = re.sub(r'\s+', ' ', t)

    # Remove balanced parenthetical content
    t = re.sub(r'\([^)]*\)', ' ', t)

    # Remove stray parentheses entirely
    t = t.replace('(', ' ').replace(')', ' ')

    # Collapse spaces again
    t = re.sub(r'\s+', ' ', t)

    # Remove trailing/leading punctuation
    t = t.strip('.,;:!?[]{}<>\\/"\' ')

    # Remove solitary characters (artifacts like orphaned 'n' after colon)
    # Only keep meaningful tokens
    t = re.sub(r'\s+', ' ', t)
    t = t.strip()
    return t


def normalize_tokens(tokens):
    """Normalize and filter a list of tokens."""
    result = []
    for t in tokens:
        t = normalize_token(t)
        if not t or len(t) < 2:
            continue
        # Drop contamination artifacts from Categories data leaking into keywords
        low = t.lower()
        if 'chevron_right' in low or 'citation topics' in low:
            continue
        # Drop tokens that are mostly numeric or special chars
        if re.match(r'^[\d.\-+]+$', t):
            continue
        result.append(t)
    return result

def load_keyword_dictionary(path):
    """Load the explicit keyword dictionary from disk."""
    if not os.path.exists(path):
        print(f"  [WARN] Dictionary not found: {path}")
        return set()
    with open(path, 'r', encoding='utf-8') as f:
        vocab = set(line.strip().lower() for line in f if line.strip())
    print(f"  Loaded {len(vocab)} entries from keyword dictionary")
    return vocab




def is_suspicious_token(token, dict_set):
    """Check if a token likely needs re-splitting using the explicit dictionary.

    A token is suspicious if it's long and not a known word/phrase.
    """
    token_low = token.lower()

    if ' ' in token:
        # Multi-word: suspicious if any single word is a long unknown run
        for w in token.split():
            w_clean = w.strip().rstrip('.,;:!?')
            if len(w_clean) > 10 and w_clean.lower() not in dict_set:
                return True
        return False
    else:
        # Spaceless: suspicious if long and not in dictionary
        if len(token) <= 9:
            return False
        return token_low not in dict_set


def greedy_split_token(token, dict_set):
    """DP-split a concatenated token against the explicit dictionary.

    For space-containing tokens, splits on spaces first, DP-splits each
    long word, then merges adjacent words back into known phrases.
    """
    if not token or len(token) < 3:
        return [token] if token else []

    if ' ' in token:
        # Split on spaces, DP-split any long unknown word
        parts = []
        for w in token.split():
            w_clean = w.strip().rstrip('.,;:!?')
            if len(w_clean) > 10 and w_clean.lower() not in dict_set:
                split_parts = _dp_split(w_clean, dict_set)
                parts.extend(split_parts)
            elif w_clean:
                parts.append(w_clean.lower())
        # Merge adjacent parts into known phrases
        result = _merge_phrases(parts, dict_set)
        if result == [token.lower()]:
            return [token]
        return result
    else:
        return _dp_split(token, dict_set)


def _merge_phrases(parts, vocab):
    """Merge adjacent words into known multi-word phrases from vocabulary."""
    if len(parts) <= 1:
        return parts
    result = []
    i = 0
    while i < len(parts):
        best = parts[i]
        best_len = 1
        for j in range(i + 1, min(i + 6, len(parts) + 1)):
            candidate = ' '.join(parts[i:j])
            if candidate in vocab:
                best = candidate
                best_len = j - i
        result.append(best)
        i += best_len
    return result


def _dp_split(token, vocab):
    """DP-split a spaceless token against vocabulary. Returns list of words."""
    token_low = token.lower()
    n = len(token_low)
    min_len = 4
    max_len = min(n, 25)

    # A token >14 chars is almost certainly concatenated, not a real word
    if token_low in vocab and len(token_low) <= 14:
        return [token_low]

    dp = [None] * (n + 1)
    dp[0] = []

    for i in range(n):
        if dp[i] is None:
            continue
        for j in range(i + 3, min(n, i + max_len) + 1):
            # Only allow 3-char match for the very last segment of the token
            if j - i == 3 and j != n:
                continue
            sub = token_low[i:j]
            if sub in vocab:
                candidate = dp[i] + [sub]
                if dp[j] is None or len(candidate) < len(dp[j]):
                    dp[j] = candidate

    if dp[n] is not None and len(dp[n]) >= 2:
        return dp[n]
    return [token_low]


# ── Research Field Extraction ───────────────────────────────────────────────

def extract_research_areas(cat_str):
    """Extract WoS Research Area names from Categories/Classification string."""
    if not isinstance(cat_str, str) or not cat_str.strip():
        return []

    s = cat_str.strip()

    # Normalize: WoS data often uses hyphenated "Other Topics" (e.g.
    # "Science & Technology - Other Topics") but the official Zendesk
    # list uses no hyphen ("Science & Technology Other Topics").
    # Handle both " - Other Topics" and "-Other Topics" variants.
    s = s.replace(" - Other Topics", " Other Topics").replace("-Other Topics", "Other Topics")
    # "Life Sciences & Biomedicine Other Topics" needs to match
    # "Life Sciences Biomedicine Other Topics" in the official list
    s = s.replace("Life Sciences & Biomedicine Other Topics", "Life Sciences Biomedicine Other Topics")

    # Find "Research Areas" section
    ra_idx = s.find("Research Areas")
    if ra_idx == -1:
        return _greedy_match_areas(s)

    start = ra_idx + len("Research Areas")
    ct_idx = s.find(" Citation Topics", start)
    if ct_idx != -1:
        areas_raw = s[start:ct_idx].strip()
    else:
        areas_raw = s[start:].strip()

    if not areas_raw:
        return []

    return _greedy_match_areas(areas_raw)


def _greedy_match_areas(text):
    """Greedy longest-match against known Research Area names."""
    sorted_areas = sorted(KNOWN_RESEARCH_AREAS, key=len, reverse=True)
    result = []
    i = 0
    n = len(text)

    while i < n:
        # Skip whitespace, newlines, and special chars
        if text[i].isspace() or text[i] in '\n\r\t':
            i += 1
            continue

        matched = False
        rem = text[i:]
        for area in sorted_areas:
            # Case-insensitive prefix match
            if rem.lower().startswith(area.lower()):
                # Verify boundary: next char should be space, uppercase start of next area, or end
                next_pos = i + len(area)
                boundary_ok = (next_pos >= n
                               or text[next_pos].isspace()
                               or text[next_pos] in '\n\r'
                               or (text[next_pos].isupper()
                                   and next_pos + 1 < n
                                   and text[next_pos + 1].islower())
                               or not text[next_pos].isalpha())
                if boundary_ok:
                    result.append(area)
                    i = next_pos
                    matched = True
                    break

        if not matched:
            i += 1

    return result


# ── Data Loading ────────────────────────────────────────────────────────────

def load_and_parse_sheet(filepath, sheet_name, year, errors):
    """Load one sheet, parse all rows, return list of record dicts."""
    records = []

    # Use openpyxl directly for more control
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  [WARN] Sheet '{sheet_name}' not found in {year}")
        wb.close()
        return records

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows()
    header_row = [cell.value for cell in next(rows_iter)]
    col_map = build_column_index(header_row, sheet_name, year)
    required = ["Year", "Keywords", "country", "Categories"]
    missing = [k for k in required if k not in col_map]
    if missing:
        print(f"  [ERROR] {year} {sheet_name}: Cannot find columns: {missing}")
        wb.close()
        return records

    for row_idx, row in enumerate(rows_iter, start=2):
        values = [cell.value for cell in row]

        def get_val(key):
            idx = col_map.get(key)
            if idx is None or idx >= len(values):
                return None
            return values[idx]

        year_val = get_val("Year")
        country_val = get_val("country")
        keywords_raw = get_val("Keywords")
        categories_raw = get_val("Categories")
        country_code_val = get_val("country_code")

        # Normalize
        if year_val is not None:
            try:
                year_val = int(year_val)
            except (ValueError, TypeError):
                year_val = year

        country_str = str(country_val).strip() if country_val is not None and str(country_val).strip().lower() not in ('none', 'nan', '') else ""
        keywords_str = str(keywords_raw).strip() if keywords_raw is not None and str(keywords_raw).strip().lower() not in ('none', 'nan', '') else ""
        categories_str = str(categories_raw).strip() if categories_raw is not None and str(categories_raw).strip().lower() not in ('none', 'nan', '') else ""
        cc_str = str(country_code_val).strip() if country_code_val is not None and str(country_code_val).strip().lower() not in ('none', 'nan', '') else ""

        # Error checking
        has_error = False
        if not country_str:
            errors.append(ErrorRecord(
                source_year=year, source_sheet=sheet_name, source_row=row_idx,
                reason="missing_country", raw_keywords=keywords_str,
                raw_categories=categories_str, raw_country=""
            ))
            has_error = True
        if not keywords_str:
            errors.append(ErrorRecord(
                source_year=year, source_sheet=sheet_name, source_row=row_idx,
                reason="missing_keywords", raw_keywords="",
                raw_categories=categories_str, raw_country=country_str
            ))
            has_error = True
        # Detect contaminated Keywords: some records have Categories data
        # copied into the Keywords column (starts with "Research Areas"
        # or contains Citation Topics / chevron_right).
        if keywords_str and not has_error:
            if (keywords_str.startswith("Research Areas")
                    or "Citation Topics" in keywords_str
                    or "chevron_right" in keywords_str):
                errors.append(ErrorRecord(
                    source_year=year, source_sheet=sheet_name, source_row=row_idx,
                    reason="contaminated_keywords", raw_keywords=keywords_str,
                    raw_categories=categories_str, raw_country=country_str
                ))
                has_error = True
        # Detect missing / invalid research field
        field_missing = False
        if not categories_str:
            field_missing = True
        elif not ("Research Areas" in categories_str
                  or "Citation Topics" in categories_str
                  or "chevron_right" in categories_str):
            # Categories field exists but has no real area data
            # (e.g. just "English" or "Spanish")
            field_missing = True
        if field_missing:
            errors.append(ErrorRecord(
                source_year=year, source_sheet=sheet_name, source_row=row_idx,
                reason="missing_field", raw_keywords=keywords_str,
                raw_categories=categories_str, raw_country=country_str
            ))
            has_error = True

        if has_error:
            continue

        # Parse keywords
        author_part, kwplus_part = split_author_vs_kwplus(keywords_str)
        author_kw = split_author_keywords(author_part) if author_part else []
        kwplus_kw = split_keywords_plus(kwplus_part) if kwplus_part else []

        # Parse research areas
        research_areas = extract_research_areas(categories_str)

        records.append({
            "year": year_val,
            "sheet_name": sheet_name,
            "source_row": row_idx,
            "country": country_str,
            "country_code": cc_str,
            "raw_keywords": keywords_str,
            "raw_categories": categories_str,
            "author_keywords": normalize_tokens(author_kw),
            "kwplus_keywords": normalize_tokens(kwplus_kw),
            "research_areas": research_areas if research_areas else ["Unknown"],
            "broad_categories": list(dict.fromkeys(
                RESEARCH_AREA_TO_BROAD.get(a, "Unknown") for a in (research_areas or ["Unknown"])
            )),
        })

    wb.close()
    return records


# ── Two-Pass Refinement ─────────────────────────────────────────────────────

def pass1_load_all():
    """Pass 1: Load all data, parse with camelCase only."""
    all_records = []
    all_errors = []

    for year in YEARS:
        filepath = os.path.join(INPUT_DIR, f"四张sheet_国家_颜色{year}.xlsx")
        if not os.path.exists(filepath):
            print(f"[WARN] File not found: {filepath}")
            continue

        print(f"Pass 1 - Loading {year}...")
        for sheet_name in SHEET_NAMES:
            records = load_and_parse_sheet(filepath, sheet_name, year, all_errors)
            all_records.extend(records)

    return all_records, all_errors


def pass2_refine(all_records):
    """Pass 2: Re-split suspicious tokens using the file-based keyword dictionary."""

    # Load the explicit dictionary (file-based, human-editable)
    dict_set = load_keyword_dictionary(KEYWORD_DICT_PATH)
    if not dict_set:
        print("  [WARN] No file dictionary, skipping Pass 2 refinement")
        for rec in all_records:
            rec["all_keywords"] = normalize_tokens(rec["author_keywords"])
        return

    total_tokens = 0
    split_count = 0
    for rec in all_records:
        refined_author = []
        for token in rec["author_keywords"]:
            total_tokens += 1
            if is_suspicious_token(token, dict_set):
                split_result = greedy_split_token(token, dict_set)
                refined_author.extend(split_result)
                if len(split_result) > 1:
                    split_count += 1
            else:
                refined_author.append(token)
        rec["author_keywords"] = refined_author

        # Only use Author Keywords for final output
        rec["all_keywords"] = normalize_tokens(rec["author_keywords"])

    print(f"  Refined {total_tokens} tokens, split {split_count} suspicious tokens")


# ── Aggregation ─────────────────────────────────────────────────────────────

def aggregate(all_records):
    """Group by (broad_category, country, year) and compute keyword frequencies."""
    # Structure: {broad_category: {country: {year: Counter}}}
    agg = defaultdict(lambda: defaultdict(lambda: defaultdict(Counter)))

    for rec in all_records:
        for cat in rec["broad_categories"]:
            country = rec["country"]
            year = rec["year"]
            for kw in rec["all_keywords"]:
                agg[cat][country][year][kw] += 1

    return agg


# ── Output Generation ───────────────────────────────────────────────────────

def write_txt_output(agg):
    """Write a single formatted txt file with all research fields."""
    filepath = os.path.join(OUTPUT_DIR, "词频统计结果.txt")
    lines = []

    for area in sorted(agg.keys()):
        countries = agg[area]
        lines.append(f"【{area}】")
        lines.append("")

        for country in sorted(countries.keys()):
            years = countries[country]
            lines.append(f"({country})")

            for year in sorted(years.keys()):
                counter = years[year]
                lines.append(f"({year})")
                for kw, count in sorted(counter.items(), key=lambda x: (-x[1], x[0].lower())):
                    lines.append(f"{kw}: {count}")
                lines.append("")

    with open(filepath, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))

    total_kw = sum(len(agg[a][c][y]) for a in agg for c in agg[a] for y in agg[a][c])
    print(f"  Wrote: 词频统计结果.txt ({len(agg)} fields, {total_kw} unique keyword entries)")


def write_mapping_excel(all_records):
    """Write raw data mapping table."""
    filepath = os.path.join(OUTPUT_DIR, "原始数据映射表.xlsx")
    rows = []
    for i, rec in enumerate(all_records):
        rows.append({
            "record_id": i + 1,
            # 记录位置
            "source_year": rec["year"],
            "source_sheet": rec["sheet_name"],
            "source_row": rec["source_row"],
            # 国家：代码 → 全称
            "country": rec["country"],
            "country_code": rec["country_code"],
            # 领域：原始 → 切分 → 大类
            "raw_categories": rec["raw_categories"],
            "research_areas": " | ".join(rec["research_areas"]),
            "broad_categories": " | ".join(rec["broad_categories"]),
            # 关键词：原始 → 作者关键词
            "raw_keywords": rec["raw_keywords"],
            "author_keywords": " | ".join(rec["author_keywords"]),
        })

    df = pd.DataFrame(rows)
    df.to_excel(filepath, index=False, engine="openpyxl")
    print(f"  Wrote: 原始数据映射表.xlsx ({len(rows)} records)")


def write_error_excel(errors):
    """Write missing/error data records."""
    filepath = os.path.join(OUTPUT_DIR, "缺失数据记录.xlsx")
    rows = [{
        "source_year": e.source_year,
        "source_sheet": e.source_sheet,
        "source_row": e.source_row,
        "reason": e.reason,
        "raw_keywords": e.raw_keywords,
        "raw_categories": e.raw_categories,
        "raw_country": e.raw_country,
    } for e in errors]

    df = pd.DataFrame(rows)
    df.to_excel(filepath, index=False, engine="openpyxl")
    print(f"  Wrote: 缺失数据记录.xlsx ({len(rows)} error records)")


# ── Verification ────────────────────────────────────────────────────────────

def verify(all_records, all_errors, agg):
    """Print verification summary."""
    print("\n" + "=" * 60)
    print("VERIFICATION SUMMARY")
    print("=" * 60)

    total_records = len(all_records)
    total_errors = len(all_errors)
    print(f"Total valid records: {total_records}")
    print(f"Total error records: {total_errors}")

    # Per-year breakdown
    year_counts = Counter(r["year"] for r in all_records)
    year_error_counts = Counter(e.source_year for e in all_errors)
    print("\nPer-year breakdown:")
    for y in sorted(YEARS):
        print(f"  {y}: {year_counts.get(y, 0)} valid + {year_error_counts.get(y, 0)} errors")

    # Per-sheet breakdown
    sheet_counts = Counter(r["sheet_name"] for r in all_records)
    print("\nPer-sheet breakdown:")
    for sn in SHEET_NAMES:
        print(f"  {sn}: {sheet_counts.get(sn, 0)} records")

    # Broad category distribution
    cat_counts = Counter()
    for rec in all_records:
        for cat in rec["broad_categories"]:
            cat_counts[cat] += 1
    print("\nBroad category distribution:")
    for cat, count in cat_counts.most_common():
        print(f"  {cat}: {count}")

    # Unique keyword count
    all_kw = set()
    for rec in all_records:
        for kw in rec["all_keywords"]:
            all_kw.add(kw.lower())
    print(f"\nUnique keywords (case-insensitive): {len(all_kw)}")

    # Total frequency entries
    total_entries = sum(
        len(agg[area][country][year])
        for area in agg for country in agg[area] for year in agg[area][country]
    )
    print(f"Total (area, country, year, keyword) frequency entries: {total_entries}")


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    global INPUT_DIR, OUTPUT_DIR, YEARS

    parser = argparse.ArgumentParser(description="Process bibliometric data")
    parser.add_argument("--input-dir", default=INPUT_DIR, help="Directory containing Excel files")
    parser.add_argument("--output-dir", default=OUTPUT_DIR, help="Output directory")
    parser.add_argument("--years", type=int, nargs="+", default=YEARS, help="Years to process")
    args = parser.parse_args()

    INPUT_DIR = args.input_dir
    OUTPUT_DIR = args.output_dir
    YEARS = args.years

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print("=" * 60)
    print("BIBLIOMETRIC DATA PROCESSING PIPELINE")
    print(f"Input: {INPUT_DIR}")
    print(f"Output: {OUTPUT_DIR}")
    print(f"Years: {YEARS}")
    print("=" * 60)

    # Phase 1: Pass 1 - Load all data
    print("\n── Phase 1: Pass 1 - Loading and initial parsing ──")
    all_records, all_errors = pass1_load_all()
    print(f"  Loaded {len(all_records)} valid records, {len(all_errors)} errors")

    # Phase 2: Load explicit keyword dictionary
    print("\n── Phase 2: Loading keyword dictionary ──")

    # Phase 3: Pass 2 - Refine with dictionary
    print("\n── Phase 3: Pass 2 - Refining keyword parsing ──")
    pass2_refine(all_records)

    # Print some refined examples for spot-check
    print("\n  Sample refined keywords:")
    for rec in all_records[:5]:
        print(f"    Row {rec['source_row']} ({rec['sheet_name']}): "
              f"author={rec['author_keywords'][:3]}..., "
              f"kwplus={rec['kwplus_keywords'][:3]}...")

    # Phase 4: Aggregate
    print("\n── Phase 4: Aggregating by field/country/year ──")
    agg = aggregate(all_records)
    print(f"  {len(agg)} research areas")

    # Phase 5: Generate output
    print("\n── Phase 5: Writing output files ──")
    write_txt_output(agg)
    write_mapping_excel(all_records)
    write_error_excel(all_errors)

    # Verify
    verify(all_records, all_errors, agg)

    print("\nDone!")


if __name__ == "__main__":
    main()
