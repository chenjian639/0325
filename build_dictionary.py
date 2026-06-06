"""
Build an explicit keyword dictionary from the current mapping table.
Outputs: keyword_dictionary.txt (one entry per line, lowercase)

Entries include:
- Single words (freq >= 3 and either short or known NLTK word)
- Multi-word phrases (freq >= 3, all component words short)
- Manually curated domain terms
- NLTK words filtered: >= 4 chars, all lowercase alpha, no numbers/symbols
"""

import re
import os
from collections import Counter
import openpyxl

# ── Load NLTK English word list (filtered) ──

_NLTK_WORDS = set()
try:
    import nltk
    nltk.download('words', quiet=True)
    for w in nltk.corpus.words.words():
        wlow = w.lower()
        if not wlow.isalpha():
            continue
        if not wlow.isascii():
            continue
        if 4 <= len(wlow) <= 20:
            _NLTK_WORDS.add(wlow)
    print("Loaded %d NLTK words (filtered: alpha-only, 4-20 chars)" % len(_NLTK_WORDS))
except Exception:
    print("NLTK not available, skipping English word list")

# ── Load all current tokens from mapping table ──

print("Loading mapping table...")
wb = openpyxl.load_workbook('output/原始数据映射表.xlsx', read_only=True)
ws = wb.active
header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
ak_col = header.index('author_keywords')

all_tokens = Counter()
for row in ws.iter_rows(min_row=2, values_only=True):
    ak = str(row[ak_col]) if row[ak_col] else ''
    for t in ak.split(' | '):
        t = t.strip()
        if t and len(t) >= 2:
            all_tokens[t] += 1
wb.close()

print("Total tokens: %d, unique: %d" % (sum(all_tokens.values()), len(all_tokens)))

# ── Load existing permanent dictionary as base ──

dictionary = set()
dict_path = 'output/keyword_dictionary.txt'
if os.path.exists(dict_path):
    with open(dict_path) as f:
        for line in f:
            line = line.strip()
            if line:
                dictionary.add(line.lower())
    print("Loaded %d entries from existing dictionary" % len(dictionary))
else:
    print("No existing dictionary found, building from scratch")

# ── Discover new words from current tokens ──

new_from_tokens = 0
for token, freq in all_tokens.items():
    low = token.lower()

    if '-' in token:
        # Keep hyphenated words
        if freq >= 3:
            dictionary.add(low)
    elif ' ' in token:
        # Multi-word phrase: keep if all words are short (not concatenated)
        words = token.split()
        lens = [len(w) for w in words if w.isalpha()]
        max_len = max(lens) if lens else 0
        if max_len <= 11 and freq >= 3:
            dictionary.add(low)
    else:
        # Single word: only keep if clearly NOT concatenated
        # Token ≤ 8 chars: safe to auto-add (freq ≥ 3)
        # Token 9-14 chars: must be in NLTK to verify it's a real word
        # Token > 14 chars: never add spaceless (likely concatenated)
        if freq >= 3:
            if len(token) <= 8:
                dictionary.add(low)
            elif len(token) <= 14 and low in _NLTK_WORDS:
                dictionary.add(low)

# Add NLTK words — only those that actually appear in our token data
# This prevents obscure NLTK-only words from bloating the dictionary
nltk_in_data = 0
for w in _NLTK_WORDS:
    # Only add if this NLTK word appears in our actual data (any freq)
    if w in all_tokens or w.lower() in all_tokens:
        dictionary.add(w)
        nltk_in_data += 1
print("NLTK words appearing in data: %d" % nltk_in_data)

# Add common NLTK words (length 4-14) directly — these are ALL verified
# English words, not fragments. This gives DP a rich vocabulary for
# splitting concatenated tokens.
nltk_direct = 0
for w in _NLTK_WORDS:
    wlen = len(w)
    if 4 <= wlen <= 14 and w not in dictionary:
        dictionary.add(w)
        nltk_direct += 1
print("NLTK direct additions: %d" % nltk_direct)

print("Base dictionary: %d entries" % len(dictionary))

# ── DP split for concatenated tokens ──

def dp_split(token, vocab):
    """DP-split a spaceless token. Returns list of words."""
    token_low = token.lower()
    n = len(token_low)
    min_len = 4
    max_len = min(n, 25)

    if token_low in vocab and len(token_low) <= 14:
        return [token_low]

    dp = [None] * (n + 1)
    dp[0] = []

    for i in range(n):
        if dp[i] is None:
            continue
        for j in range(i + min_len, min(n, i + max_len) + 1):
            sub = token_low[i:j]
            if sub in vocab:
                candidate = dp[i] + [sub]
                if dp[j] is None or len(candidate) < len(dp[j]):
                    dp[j] = candidate

    if dp[n] is not None and len(dp[n]) >= 2:
        return dp[n]
    return [token_low]

# ── Split concatenated spaceless tokens ──

concat_tokens = []
for token, freq in all_tokens.items():
    low = token.lower()
    if ' ' in token or '-' in token:
        continue
    if len(token) <= 9:
        continue
    if low in dictionary:
        continue
    if low in _NLTK_WORDS:
        continue
    concat_tokens.append((token, freq))

print("Concatenated tokens to split: %d" % len(concat_tokens))

split_count = 0
new_from_splits = set()
for token, freq in concat_tokens:
    result = dp_split(token, dictionary)
    if len(result) >= 2:
        split_count += 1
        for w in result:
            new_from_splits.add(w)
            dictionary.add(w)

print("Successfully split: %d / %d" % (split_count, len(concat_tokens)))

# ── Phrase merging helper ──

def _merge_phrases(parts, vocab):
    """Merge adjacent parts into known multi-word phrases from vocab."""
    if len(parts) <= 1:
        return parts
    result = []
    i = 0
    while i < len(parts):
        best = parts[i]
        best_len = 1
        for j in range(i + 1, min(i + 6, len(parts) + 1)):
            candidate = ' '.join(parts[i:j])
            if candidate in vocab:
                best = candidate
                best_len = j - i
        result.append(best)
        i += best_len
    return result

# ── Split space-containing tokens with long words ──

phrase_count = 0
for token, freq in all_tokens.items():
    if ' ' not in token:
        continue
    words = token.split()
    needs_split = False
    for w in words:
        w_clean = w.strip().rstrip('.,;:!?')
        if (len(w_clean) > 9 and w_clean.isalpha()
                and w_clean.lower() not in _NLTK_WORDS
                and w_clean.lower() not in dictionary):
            needs_split = True
            break
    if needs_split:
        new_parts = []
        for w in words:
            w_clean = w.strip().rstrip('.,;:!?')
            if len(w_clean) > 14 and w_clean.isalpha() and w_clean.lower() not in dictionary:
                parts = dp_split(w_clean, dictionary)
                new_parts.extend(parts)
            elif w_clean:
                new_parts.append(w_clean.lower())
        # After splitting, try to merge adjacent words into known phrases
        merged = _merge_phrases(new_parts, dictionary)
        if merged != [w.lower() for w in words]:
            phrase_count += 1
            for p in merged:
                if p not in dictionary and ' ' not in p and len(p) >= 3:
                    new_from_splits.add(p)
                    dictionary.add(p)
                elif ' ' in p and p not in dictionary:
                    dictionary.add(p)

print("Space-containing tokens restructured: %d" % phrase_count)

# ── Manual domain additions ──

domain_terms = {
    # Core ontology / knowledge organization
    'ontology', 'ontologies', 'ontological',
    'semantic', 'semantics', 'semantic web', 'semantic interoperability',
    'knowledge', 'knowledge management', 'knowledge representation',
    'knowledge engineering', 'knowledge graph', 'knowledge organization',
    'knowledge discovery', 'knowledge base', 'knowledge bases',
    'knowledge sharing', 'knowledge acquisition',
    'metadata', 'linked data', 'linked open data', 'open data',
    'thesaurus', 'thesauri', 'taxonomy', 'taxonomies',
    'classification', 'categorization', 'indexing',
    'information retrieval', 'information extraction',
    'information science', 'library science',
    'information system', 'information systems',
    'natural language processing', 'nlp',
    'machine learning', 'deep learning', 'artificial intelligence',
    'data mining', 'text mining', 'web mining', 'process mining',
    'reasoning', 'inference', 'description logic', 'description logics',
    'owl', 'rdf', 'rdfs', 'sparql', 'xml', 'xml schema',
    'web ontology language', 'semantic web rule language', 'swrl',
    'conceptual model', 'conceptual modeling', 'conceptual modelling',
    'domain ontology', 'application ontology', 'reference ontology',
    'upper ontology', 'foundational ontology', 'core ontology',
    'ontology alignment', 'ontology matching', 'ontology mapping',
    'ontology merging', 'ontology integration',
    'ontology engineering', 'ontology development',
    'ontology learning', 'ontology evaluation', 'ontology evolution',
    'ontology modularization', 'ontology design pattern',
    'ontology visualization', 'ontology debugging',
    'ontology-based', 'ontology-driven', 'ontology population',
    'database', 'databases', 'relational database', 'relational databases',
    'expert system', 'expert systems',
    'decision support', 'decision support system',
    'recommender system', 'recommendation system',
    'query expansion', 'query reformulation',
    'knowledge organization system', 'knowledge organization systems',
    'controlled vocabulary', 'controlled vocabularies',
    # Multi-word phrases that should stay together
    'semantic web', 'linked data', 'linked open data',
    'web ontology language', 'semantic interoperability',
    'information retrieval', 'information extraction',
    'knowledge management', 'knowledge representation',
    'knowledge engineering', 'knowledge graph',
    'natural language processing', 'machine learning',
    'deep learning', 'artificial intelligence',
    'data mining', 'text mining', 'big data',
    'description logic', 'description logics',
    'ontology alignment', 'ontology matching', 'ontology mapping',
    'ontology engineering', 'ontology learning',
    'decision support', 'recommender system',
    'query expansion', 'case-based reasoning',
    'expert system', 'expert systems',
    'social network', 'social media',
    'software engineering', 'requirements engineering',
    'project management', 'risk management', 'change management',
    'electronic health record', 'electronic health records',
    'cultural heritage', 'digital library', 'digital libraries',
    'internet of things', 'iot',
    'web service', 'web services',
    'cloud computing', 'mobile computing',
    'neural network', 'neural networks',
    'support vector machine', 'random forest',
    'geographic information', 'geographic information system', 'gis',
    'computer science', 'information science',
    'life sciences', 'social sciences', 'physical sciences',
    'research area', 'research areas',
    # Domain-specific multi-word phrases from the corpus
    'olympic games', 'world war', 'latin america',
    'united states', 'united kingdom', 'south africa',
    'new zealand', 'south korea', 'saudi arabia',
    'sri lanka', 'czech republic',
    '20th century', '21st century', '19th century',
    'armed forces', 'public health', 'primary care',
    'supply chain', 'value chain',
    'user interface', 'user experience',
    'open source', 'open access',
    'real time', 'real-time',
    'state of the art', 'state-of-the-art',
    'cost benefit', 'cost effectiveness',
    'life cycle', 'lifecycle',
    'best practice', 'best practices',
    'case study', 'case studies', 'case report',
    'literature review', 'systematic review',
    # Words that commonly appear concatenated
    'semantic', 'ontological', 'conceptual', 'formal', 'informal',
    'automatic', 'automated', 'manual', 'semi-automatic',
    'structural', 'functional', 'behavioral', 'dynamic', 'static',
    'hybrid', 'fuzzy', 'probabilistic', 'statistical',
    'linguistic', 'lexical', 'syntactic', 'pragmatic', 'semantic',
    'logical', 'rule-based', 'constraint-based', 'data-driven',
    'instance', 'class', 'property', 'relation', 'relationship',
    'concept', 'term', 'entity', 'attribute', 'value',
    'individual', 'axiom', 'fact', 'rule', 'constraint',
    'network', 'graph', 'tree', 'hierarchy', 'lattice',
    'pattern', 'template', 'schema', 'model', 'framework',
    'user', 'agent', 'actor', 'stakeholder', 'provider',
    'organization', 'enterprise', 'company', 'industry', 'sector',
    'health', 'medical', 'clinical', 'pharmaceutical', 'biological',
    'education', 'learning', 'teaching', 'training', 'instruction',
    'government', 'public', 'policy', 'law', 'regulation',
    'environment', 'energy', 'transport', 'agriculture', 'food',
    'science', 'scientific', 'research', 'innovation', 'technology',
    'culture', 'history', 'language', 'literature', 'art',
    'economy', 'economics', 'business', 'finance', 'market',
    'society', 'social', 'political', 'legal', 'ethical',
    'based', 'driven', 'oriented', 'centric', 'specific',
    'related', 'general', 'common', 'shared', 'distributed',
    'evaluation', 'assessment', 'validation', 'verification',
    'generation', 'recognition', 'segmentation', 'classification',
    'prediction', 'optimization', 'simulation', 'emulation',
    'acquisition', 'representation', 'presentation', 'publication',
    'organization', 'administration', 'communication', 'collaboration',
    'identification', 'authentication', 'authorization', 'certification',
    'configuration', 'customization', 'personalization', 'adaptation',
    'transformation', 'normalization', 'standardization', 'harmonization',
    'composition', 'decomposition', 'aggregation', 'disaggregation',
    'association', 'correlation', 'comparison', 'contrast',
    'detection', 'correction', 'prevention', 'protection',
    'monitoring', 'tracking', 'logging', 'reporting', 'alerting',
    'planning', 'scheduling', 'allocation', 'assignment',
    'negotiation', 'mediation', 'coordination', 'cooperation',
    'interaction', 'interconnection', 'interoperability',
    'interpretation', 'explanation', 'description', 'prescription',
    'construction', 'reconstruction', 'deconstruction',
    'inspection', 'examination', 'investigation', 'exploration',
    'demonstration', 'illustration', 'visualization', 'animation',
    'recommendation', 'personalization', 'customization',
    'implementation', 'deployment', 'operation', 'maintenance',
    'integration', 'migration', 'conversion', 'transformation',
    'annotation', 'tagging', 'labeling', 'markup',
    'mapping', 'matching', 'alignment', 'merging', 'linking',
    'connecting', 'combining', 'splitting', 'filtering',
    'search', 'retrieval', 'discovery', 'browsing', 'navigation',
    'query', 'querying', 'indexing', 'ranking', 'sorting',
    'extraction', 'collection', 'gathering', 'harvesting',
    'storage', 'archiving', 'preservation', 'curation',
    'modeling', 'modelling', 'designing', 'developing', 'building',
    'analyzing', 'evaluating', 'testing', 'validating',
    'processing', 'computing', 'calculating', 'reasoning',
    'comparing', 'contrasting', 'differentiating',
    'managing', 'organizing', 'administering', 'governing',
    'engineering', 'constructing', 'manufacturing', 'producing',
    'using', 'applying', 'employing', 'utilizing',
    # Countries and places (critical for splitting geo-concatenated tokens)
    'afghanistan', 'albania', 'algeria', 'andorra', 'angola',
    'argentina', 'armenia', 'australia', 'austria', 'azerbaijan',
    'bahamas', 'bahrain', 'bangladesh', 'barbados', 'belarus',
    'belgium', 'belize', 'benin', 'bhutan', 'bolivia',
    'bosnia', 'botswana', 'brazil', 'brunei', 'bulgaria',
    'burkina', 'burundi', 'cambodia', 'cameroon', 'canada',
    'chad', 'chile', 'china', 'colombia', 'congo',
    'croatia', 'cuba', 'cyprus', 'czech', 'denmark',
    'djibouti', 'dominica', 'dominican', 'ecuador', 'egypt',
    'england', 'estonia', 'ethiopia', 'fiji', 'finland',
    'france', 'gabon', 'gambia', 'georgia', 'germany',
    'ghana', 'greece', 'grenada', 'guatemala', 'guinea',
    'guyana', 'haiti', 'honduras', 'hungary', 'iceland',
    'india', 'indonesia', 'iran', 'iraq', 'ireland',
    'israel', 'italy', 'jamaica', 'japan', 'jordan',
    'kazakhstan', 'kenya', 'korea', 'kuwait', 'kyrgyzstan',
    'laos', 'latvia', 'lebanon', 'lesotho', 'liberia',
    'libya', 'lithuania', 'luxembourg', 'macedonia', 'madagascar',
    'malawi', 'malaysia', 'maldives', 'mali', 'malta',
    'mauritania', 'mauritius', 'mexico', 'moldova', 'monaco',
    'mongolia', 'montenegro', 'morocco', 'mozambique', 'myanmar',
    'namibia', 'nepal', 'netherlands', 'zealand', 'nicaragua',
    'niger', 'nigeria', 'norway', 'oman', 'pakistan',
    'palestine', 'panama', 'paraguay', 'peru', 'philippines',
    'poland', 'portugal', 'qatar', 'romania', 'russia',
    'rwanda', 'samoa', 'senegal', 'serbia', 'singapore',
    'slovakia', 'slovenia', 'somalia', 'spain', 'lanka',
    'sudan', 'suriname', 'sweden', 'switzerland', 'syria',
    'taiwan', 'tanzania', 'thailand', 'togo', 'tonga',
    'trinidad', 'tunisia', 'turkey', 'turkmenistan', 'uganda',
    'ukraine', 'emirates', 'kingdom', 'states', 'uruguay',
    'uzbekistan', 'venezuela', 'vietnam', 'wales', 'yemen',
    'zambia', 'zimbabwe', 'scotland', 'britain',
    'africa', 'america', 'asia', 'europe', 'oceania',
    'latin', 'north', 'south', 'east', 'west', 'central',
    'pacific', 'atlantic', 'mediterranean', 'caribbean',
    # Additional commonly missing words
    'forces', 'dictatorship', 'government', 'state', 'armed',
    'cali', 'commerce', 'pharmacy', 'laboratory', 'industry',
    'gymnastics', 'habitus', 'tradition', 'violence', 'gender',
    'racism', 'sexuality', 'borders', 'migration', 'doctrine',
    'architecture', 'colonial', 'decolonization', 'intercultural',
    # Missing from obscure Humanities/Social Sciences papers
    'designation', 'predicative', 'expression', 'rigidity',
    'debate', 'magazines', 'magazine', 'vessels', 'vessel',
    'biography', 'biographies', 'persons', 'person',
    'pension', 'pensions', 'work', 'working', 'historical',
    'sources', 'source', 'agencies', 'agency', 'crisis',
    'crises', 'administration', 'justice', 'advertising',
    'radio', 'consumption', 'consumer', 'consuming',
    'bolshevization', 'cellular', 'frenchification', 'feeding',
    'emotion', 'emotions', 'emotional', 'feeling', 'feelings',
    'suffering', 'suffer', 'collective', 'disorder', 'disorders',
    'boarding', 'innovation', 'innovations', 'technological',
    'workers', 'worker', 'journalists', 'journalist',
    'memory', 'memories', 'memorial', 'organization',
    'organizations', 'organizational', 'body', 'bodies',
    'mexico', 'mexican', 'narrative', 'narratives', 'narration',
    'psychogenic', 'psychology', 'psychological',
    'multiculturalism', 'multicultural', 'interseccionality',
    'intersectional', 'intersectionality', 'mapuche',
    'crime', 'crimes', 'criminal', 'youth', 'youths',
    'death', 'deaths', 'dying', 'schema', 'schemas', 'schemata',
    'andin', 'andine', 'subregion', 'subregions', 'sub',
    'churches', 'church', 'cathedral', 'career', 'careers',
    'cities', 'city', 'bolivia', 'bolivian', 'plurinational',
    'sociohistorical', 'transmodernity', 'modernity',
    'conservatism', 'conservative', 'anticomunismo',
    'anticommunism', 'communism', 'communist',
    'argentina', 'argentine', 'argentinian', 'chile', 'chilean',
    'colombia', 'colombian', 'peru', 'peruvian', 'ecuador',
    'ecuadorian', 'uruguay', 'uruguayan', 'paraguay',
    'paraguayan', 'venezuela', 'venezuelan',
    'historiography', 'historiographic', 'historiographical',
    'institutional', 'institution', 'institutions',
    'headquarters', 'imperialism', 'imperial', 'imperialist',
    'international', 'secretariat', 'translation', 'translations',
    'translating', 'ethnography', 'ethnographic', 'ethnographical',
    'ethnology', 'editing', 'editor', 'editorial', 'editions',
    'groups', 'group', 'grouping', 'parties', 'party',
    'bolshevization', 'bolshevik', 'communist', 'communism',
    'advertising', 'advertise', 'advertisement', 'advertisements',
    'radio', 'radiofonic', 'broadcasting', 'broadcast',
    'agencies', 'agency', 'crisis', 'crises', 'depression',
    'administration', 'administrative', 'administer',
    'justice', 'judicial', 'judiciary', 'court', 'courts',
    'institutional', 'institution', 'ethnography',
    'political', 'politics', 'politician', 'politicians',
    'headquarters', 'headquarter', 'secretariat',
    'imperialism', 'communist', 'international',
    'frenchification', 'frenchify', 'french',
    'feeding', 'feed', 'food', 'foods', 'cuisine',
    'bolshevization', 'cellular', 'cell', 'cells',
    'sources', 'source', '19th', '20th', 'century', 'centuries',
    'pension', 'pensions', 'work', 'working', 'worker',
    'historical', 'history', 'historian', 'historians',
    'old', 'age', 'aging', 'aged', 'retirement', 'retired',
    'war', 'wars', 'warfare', 'warlike', 'warmongering',
    'peace', 'peaceful', 'pacifism', 'pacifist',
    'conflict', 'conflicts', 'conflicting', 'conflictual',
    'disorder', 'disorders', 'fear', 'fears', 'fearful',
    'trauma', 'traumas', 'traumatic', 'traumatized',
    'hospital', 'hospitals', 'health', 'healthy', 'healthcare',
    'law', 'laws', 'legal', 'legality', 'legislation',
    'equity', 'equitable', 'inequity', 'inequality',
    'san', 'juan', 'dios', 'dioses',
    'anthropology', 'anthropological', 'anthropologist',
    'narrative', 'narratives', 'narration', 'narrator',
    'suffering', 'suffer', 'pain', 'painful',
    'psychogenic', 'psychology', 'psychological', 'psychologist',
    'religious', 'religion', 'religions', 'religiosity',
    'boarding', 'board', 'school', 'schools', 'schooling',
    'youth', 'youths', 'young', 'youngster', 'youngsters',
    'organization', 'organizations', 'organizing', 'organized',
    'collective', 'collectives', 'collectivity',
    'action', 'actions', 'activism', 'activist', 'activists',
    'hijos', 'contagio', 'contagion', 'contagious',
    'workers', 'worker', 'workforce', 'workplace',
    'emotion', 'emotions', 'emotional', 'emotionally',
    'journalists', 'journalist', 'journalism', 'journal',
    'networks', 'network', 'networking', 'networked',
    'memory', 'memories', 'memorial', 'memorialization',
    'community', 'communities', 'communal', 'communitarian',
    'multiculturalism', 'multicultural', 'multiethnic',
    'tradition', 'traditions', 'traditional', 'traditionalism',
    'violence', 'violent', 'nonviolence', 'nonviolent',
    'gender', 'genders', 'gendered', 'gendering',
    'interseccionality', 'intersectional', 'intersection',
    'intersections', 'intersectionality',
    'racism', 'racist', 'racial', 'racialization', 'racialized',
    'crime', 'crimes', 'criminal', 'criminality', 'criminally',
    'black', 'blacks', 'blackness', 'afro', 'african',
    'death', 'deaths', 'dying', 'dead', 'deadly',
    'schema', 'schemas', 'schemata', 'schematic',
    'borders', 'border', 'bordering', 'borderland',
    'sexuality', 'sexualities', 'sexual', 'sexuality',
    'andin', 'andine', 'andes', 'andean',
    'subregion', 'subregions', 'subregional',
    'churches', 'church', 'ecclesiastical', 'ecclesiastic',
    'architecture', 'architectural', 'architect', 'architects',
    'doctrine', 'doctrines', 'doctrinal', 'doctrinally',
    'colonial', 'colony', 'colonies', 'colonialism', 'colonization',
    'career', 'careers', 'professional', 'profession',
    'cities', 'city', 'urban', 'metropolitan', 'municipal',
    'bolivia', 'bolivian', 'bolivians',
    'decolonization', 'decolonize', 'decolonial', 'decolonizing',
    'plurinational', 'plurinationality', 'multinational',
    'intercultural', 'interculturalism', 'interculturality',
    'sociohistorical', 'socio', 'historical', 'transmodernity',
    'modernity', 'modern', 'modernist', 'modernization',
    'conservatism', 'conservative', 'conservatives',
    'anticomunismo', 'anticommunism', 'communism', 'communist',
    'chile', 'chilean', 'chileans',
    'habitus', 'habit', 'habits', 'habituation', 'habitual',
    'exile', 'exiles', 'exiled', 'exilic',
    'multiculturalismtraditionviolencegender',
    'workersemotions', 'warmedicineemotion', 'disorderfeartrauma',
    'vesselsbiographypersons', 'historydebatemagazines',
    'pensionworkhistorical', 'argentinaadministrationjustice',
    'argentinaadvertisingconsumptionradio', 'frenchificationfeeding',
    'bolshevizationcellular', 'agenciescrisis', 'sources19th',
    'designationpredicative', 'expressionrigidityontology',
    'latin americascientific', 'conservatismpolitical',
    'Thesaurus:', 'Author:', 'Authors:',
    # French / Spanish academic terms (for Humanities papers)
    'autochtones', 'evangelisation', 'indigenisation', 'catechisme',
    'visuel', 'ontologie', 'metissage', 'religieux', 'langue',
    'innue', 'echelles', 'catholiques', 'ethnography', 'ethnographic',
    'fieldwork', 'racism', 'racial', 'immigration', 'immigrant',
    'aspectuality', 'intermediality', 'virtuality', 'avatar',
    'omniscient', 'performance', 'ethics', 'ethical',
}

dictionary.update(domain_terms)
print("Domain terms added: %d" % len(domain_terms))

# ── Second pass: re-split with enriched dictionary ──

split_count2 = 0
for token, freq in concat_tokens:
    result = dp_split(token, dictionary)
    if len(result) >= 2:
        split_count2 += 1
        for w in result:
            dictionary.add(w)

print("Second pass splits: %d (total: %d)" % (split_count2, split_count + split_count2))

# ── Write dictionary ──

# Only keep clean entries: alpha/spaces/hyphens, no leading/trailing junk
def is_clean_entry(s):
    if not s or len(s) < 2:
        return False
    if not any(c.isalpha() for c in s):
        return False
    if not (s[0].isalnum() and s[-1].isalnum()):
        return False
    for c in s:
        if not (c.islower() or c.isdigit() or c in (' ', '-', '_')):
            return False
    # Reject long spaceless entries: these are concatenated tokens that
    # failed to split and would block DP if kept in the dictionary
    if ' ' not in s and len(s) > 14:
        return False
    return True

clean_dict = sorted(w for w in dictionary if is_clean_entry(w))

with open('output/keyword_dictionary.txt', 'w', encoding='utf-8') as f:
    for word in clean_dict:
        f.write(word + '\n')

print("\nDictionary written to output/keyword_dictionary.txt")
print("Entries: %d (filtered from %d)" % (len(clean_dict), len(dictionary)))
print("\nDone!")
